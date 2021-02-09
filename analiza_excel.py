#! python3

from openpyxl import load_workbook
import pyperclip
import re
import os
import sys
import pprint
from colorama import init, Fore, Back, Style


def check_mc(column, worksheet, file_name):
    """Funcja sprawdza czy w opisie realizacji i nazwie pliku jest wskazana ta sama długość umowy"""
    ws = worksheet
    warnings = []  # Lista do umieszczania komuniktów
    period_desc_duration = False
    period_file_duration = False
    file = os.path.basename(file_name)

    reg_mc = re.compile(r'_(\d+)mc')  # do wyszukania okresu wariantu w nazwie pliku
    period_file = reg_mc.search(file)

    if period_file is None:
        warnings.append("W nazwie pliku nie określono długości trwania umowy.")
    else:
        period_file_duration = period_file.group(1)

    reg_mc2 = re.compile(r'[W,w]ariant.+\s(\d+)\s*mc')  # do wyszukania okresu wariantu w opisie
    period_desc = reg_mc2.search(ws[f'C{column}'].value)

    if period_desc is None:
        warnings.append("W opisie realizacji nie określono długości trwania umowy.")
    else:
        period_desc_duration = period_desc.group(1)

    if period_desc_duration and period_file_duration:
        # print("W opisie:", period_desc_duration)
        # print("W nazwie:", period_file_duration)
        if period_file_duration != period_desc_duration:
            warnings.append("Długość trwania umowy w opisie niezgodna z nazwą pliku")

    return warnings, period_file_duration, period_desc_duration


def data(row_ns, file_duration, desc_duration, worksheet, w_count):
    """Tworzy słownik z danymi dla danego numeru sprawy CAS"""
    ws = worksheet
    warning = []
    db = {}  # Słownik do którego trafią dane dla poszczególnych PSK/DLC/LD itp
    check = {}  # Wykorzystany do rozwiązywania problemu z taką samą nazwą psp dla jednego PSK/DLC/LD itp

    # ignore = [2, 6]
    for z in range(int(row_ns) + 1, w_count):  # Przelatuje wiersze od row_ns (Numer sprawy) do końca arkusza
        if ws[f'N{z}'].value:  # jeżeli komórka zawiera numer lokalizacji
            # print(ws[f'N{z}'].value)
            mc_size = 1  # Marged cell size
            for w in range(z + 1, w_count):  # Sprawdza ile wierszy składa się na scaloną komórkę z nr PSK/DLC/itp
                if ws[f'N{w}'].value is None and ws[f'S{w}'].value is not None:
                    mc_size += 1  # Marged cell size
                else:
                    break

            db[ws[f'N{z}'].value] = {}
            db[ws[f'N{z}'].value][ws[f'O{row_ns}'].value] = ws[f'O{z}'].value  # Lokalizacja
            db[ws[f'N{z}'].value][ws[f'Q{row_ns}'].value] = ws[f'Q{z}'].value  # Wycena DA lub DTWS-Capex
            db[ws[f'N{z}'].value][ws[f'R{row_ns}'].value] = ws[f'R{z}'].value  # Wycena DA lub DTWS-OPEX
            db[ws[f'N{z}'].value][ws[f'S{row_ns}'].value] = {}  # Nazwa PSP
            # Ponieważ słownik nie może mieć identycznych kluczy stąd poniższe rozwiązanie
            check.clear()  # Czyści słownik przed każdą DLC/PSK/itp
            mrc = 0
            for x in range(z, z + mc_size):  # Sprawdza wszystkie wiersze scalonej komórki
                if ws[f'S{x}'].value in db[ws[f'N{z}'].value][
                    ws[f'S{row_ns}'].value]:  # jeżeli element nazwa PSP istnieje
                    warning.append(f"[{ws[f'N{z}'].value}] : [{ws[f'S{x}'].value}] - PSP wystąpił więcej niż raz")
                    if ws[f'S{x}'].value not in check:  # i jeżeli nie ma go jeszcze w dict check
                        check[ws[f'S{x}'].value] = 1  # dodajemy go do słownika z wartością 1
                    else:  # natomiast jeżeli jest w dict check
                        check[ws[f'S{x}'].value] += 1  # podnosimy jego wartość o +1
                        # check[ws[f'S{x}'].value] = check.get(ws[f'S{x}'].value) + 1
                    psp = ws[f'S{x}'].value + '*' * check[
                        ws[f'S{x}'].value]  # Do nazwy PSP dodajemy * pomnożoną o licznik wystąpień
                else:
                    psp = ws[f'S{x}'].value

                # Poniżej zbieramy wartości dla poszczególnych psp
                db[ws[f'N{z}'].value][ws[f'S{row_ns}'].value][psp] = {}
                db[ws[f'N{z}'].value][ws[f'S{row_ns}'].value][psp][ws[f'T{row_ns}'].value] = ws[
                    f'T{x}'].value  # PSP
                db[ws[f'N{z}'].value][ws[f'S{row_ns}'].value][psp][ws[f'U{row_ns}'].value] = ws[
                    f'U{x}'].value  # Ilość
                db[ws[f'N{z}'].value][ws[f'S{row_ns}'].value][psp][ws[f'V{row_ns}'].value] = ws[
                    f'V{x}'].value  # Wartość
                db[ws[f'N{z}'].value][ws[f'S{row_ns}'].value][psp][ws[f'W{row_ns}'].value] = ws[
                    f'W{x}'].value  # Suma
                db[ws[f'N{z}'].value][ws[f'S{row_ns}'].value][psp][ws[f'X{row_ns}'].value] = ws[
                    f'X{x}'].value  # Rodzaj kosztu
                db[ws[f'N{z}'].value][ws[f'S{row_ns}'].value][psp][ws[f'Y{row_ns}'].value] = ws[
                    f'Y{x}'].value  # Grupa księgowa
                if ws[f'X{x}'].value == 'Opex':
                    if (ws[f'U{x}'].value != int(file_duration)) or (ws[f'U{x}'].value != int(desc_duration)):
                        warning.append(f"[{ws[f'N{z}'].value}] : [{psp}] - ilość: {ws[f'U{x}'].value} - "
                                       f"Porównaj z opisem i nazwą pliku")
                    else:
                        mrc += float(ws[f'V{x}'].value)
                if ws[f'X{x}'].value == 'Capex':
                    warning.append(f"[{ws[f'N{z}'].value}] : [{ws[f'S{x}'].value}] - koszt w postaci Capex")
            db[ws[f'N{z}'].value]["Koszt miesięczny sumaryczny"] = mrc

    return db, warning


def show_dict(result, warning):
    print(Fore.LIGHTWHITE_EX, Style.BRIGHT)
    for i in result.keys():
        if i in str(warning):
            print(Fore.LIGHTYELLOW_EX, Back.RED, f'{i} :', Fore.RESET, Back.RESET)
        else:
            print(Fore.LIGHTYELLOW_EX, f'{i} :', Fore.RESET)
        for j in result[i].keys():
            if not isinstance(result[i][j], dict):
                print('        ', f'{j.strip()} : {result[i][j]}')
            else:
                print('        ', j.strip(), ':', end=None)
                for k in result[i][j]:
                    print(Fore.LIGHTCYAN_EX, '        ', k, ':', Fore.RESET, end=None)
                    for l in result[i][j][k]:
                        print('        ', '        ', l, ':', result[i][j][k][l])
    print(Fore.RESET, Style.RESET_ALL)


def show_warnings(msg):
    print(Fore.LIGHTYELLOW_EX)
    print('###########')
    print('Sprawdź to:')
    print('###########')
    print(Fore.RESET)
    print()
    if msg:
        for i in msg:
            print(Back.RED + Fore.LIGHTYELLOW_EX + i)
            print()
    else:
        print(Fore.LIGHTCYAN_EX + "Brak uwag" + Fore.RESET)
    print(Fore.RESET + Back.RESET)


def check_ws(worksheet):
    ws = worksheet
    row_count = 0
    col_description = None
    col_numer_sprawy = None
    addr = str(ws['E19'].value)  # Location address - fix location in worksheet
    reg_num = re.compile(r'[A-Z](\d+)>')  # Looks for row number in string
    reg_term = re.compile(r'Termin realizacji\s(.*)')
    w = ws.column_dimensions['C'].width + ws.column_dimensions['D'].width + ws.column_dimensions['E'].width + \
        ws.column_dimensions['F'].width + ws.column_dimensions['G'].width + ws.column_dimensions['H'].width  # Check
    # for merged cell width

    for i in ws['E']:
        row_count += 1
        try:
            if addr in str(i.value):
                # print(str(i))
                col3 = int(reg_num.search(str(i)).group(1))
                # print(col3)
                ws.row_dimensions[col3].height = (abs(len(ws[f'E{col3}'].value) // (-18))) * 12.75  # round value up
                # (-18) is a number of letter that fit in one line of localisation address cell

        except ValueError:
            continue

    for v in ws['A']:

        try:
            if "Data uruchomienia" in str(v.value):
                col = reg_num.search(str(v)).group(1)  # get cell coord form object name

            elif "Opis realizacji" in str(v.value):
                col_description = reg_num.search(str(v)).group(1)  # get cell coord form object name
                match = reg_term.search(str(ws[f'C{col_description}'].value))  # search for string with period value
                s = 0
                for i in ws[f'C{col_description}'].value.splitlines():  # Count how many line need to fit inside a cell
                    # print(i, len(i))
                    if len(i) > int(w):
                        s = s + len(i) // int(w)
                    else:
                        s = s + 1
                # print(s, 'liczba linii')
                ws.row_dimensions[int(col_description)].height = 12.5 * s
                if match is not None:
                    ws[f'C{col}'] = str(match.group(1))  # Add period next to cell with 'Data uruchomienia'
                else:
                    ws[f'C{col}'] = 'ND'  # Add ND next to cell with 'Data uruchomienia'
            elif "Numer sprawy" in str(v.value):
                col_numer_sprawy = reg_num.search(str(v)).group(1)  # get cell coord form object name
                # print(w_count)
                # res, warn = data(int(col4))
        except ValueError:
            continue
    return col_description, col_numer_sprawy, row_count


def file_processing(file_name, var_count=0, variants=None):
    wb = load_workbook(file_name)
    wb.save(file_name)
    wb.close()
    wb = load_workbook(file_name)
    ws = wb.active

    ws.column_dimensions['C'].width = 8.5
    ws.row_dimensions[18].height = 25
    ws.column_dimensions['D'].width = len(ws['D18'].value)
    ws.column_dimensions['H'].width = len(ws['H18'].value)
    ws.column_dimensions['E'].width = 16

    col_desc, col_ns, w_count = check_ws(ws)
    warn_period, file_period, desc_period = check_mc(col_desc, ws, file_name)
    res, warn = data(int(col_ns), file_period, desc_period, ws, w_count)
    print()
    print(Fore.LIGHTYELLOW_EX + "Nazwa pliku: " + Fore.RESET, Fore.LIGHTCYAN_EX, os.path.basename(file_name),
          Fore.RESET)
    print(Fore.LIGHTYELLOW_EX + "Lokalizacja: " + Fore.RESET, Fore.LIGHTCYAN_EX, os.path.dirname(file_name),
          Fore.RESET)
    print()
    show_dict(res, warn)

    print()
    show_warnings(warn)
    print()
    input("Sprawdź powyższe i wciśnij Enter, aby wyświetlić opis realizacji")
    print()
    print()
    print(Fore.LIGHTYELLOW_EX + "Nazwa pliku: " + Fore.RESET, Fore.LIGHTCYAN_EX, os.path.basename(file_name),
          Fore.RESET)
    print(Fore.LIGHTYELLOW_EX + "Lokalizacja: " + Fore.RESET, Fore.LIGHTCYAN_EX, os.path.dirname(file_name),
          Fore.RESET)
    print(Fore.LIGHTYELLOW_EX)
    print("#" * 32, "Opis realizacji", '#' * 32)
    print(Fore.LIGHTCYAN_EX)
    print(ws[f'C{col_desc}'].value)
    print(Fore.LIGHTYELLOW_EX)
    print("#" * 32, "Opis realizacji", '#' * 32)
    print(Fore.RESET)
    # warn_period = check_mc(col_desc)
    show_warnings(warn_period)
    print()
    input("Sprawdź powyższe i wciśnij Enter, aby kontynuować")
    if variants:
        if var_count == len(variants):
            pyperclip.copy('\n'.join(variants) + '\n\n' + ws[f'C{col_desc}'].value)
    else:
        pyperclip.copy(ws[f'C{col_desc}'].value)
    wb.save(file_name)
    wb.close()


def variants_check(file_name):
    """Sprawdza, czy w tym samym katalogu znajdują się inne warianty dla tej samej analizy"""
    file_name_only = os.path.basename(file_name)
    service_type = re.search(r'_SDI_|_DLC_|_PSK_', file_name_only, re.IGNORECASE)
    if service_type:
        client = file_name_only[:file_name_only.find(service_type[0])]
    else:
        client = file_name_only

    print(Fore.LIGHTWHITE_EX, Style.BRIGHT)
    print("Analiza dla Klienta:", client)
    print()
    find_file_date = re.compile(r'_\d\d_\d\d_\d\d\d\d')
    file_date = find_file_date.findall(file_name_only)[0].strip('_')
    print("Data analizy:", file_date)
    print()
    all_file_list = os.listdir(os.path.dirname(file_name))

    files_variants = []
    for file in all_file_list:
        if client in file and file_date in file:
            files_variants.append(file)

    if len(files_variants) > 1:
        print('Wykryto dodatkowe warianty.')
        print()
        for var in files_variants:
            print(var)
        print()
        ask = input("Czy uwzględnić (t/n)? ")
        print()
        if ask.lower() == 't':
            print('Procesuję wszystkie warianty.')
            print()
            variants_count = len(files_variants)  # liczba wszystkich wariantów
            variants = []  # lista z opisami poszczególnych wariantów
            for var in files_variants:
                variants.append(variants_text(var, variants))  # Dodaje do listy opisy poszczególnych wariantów
                file_processing(os.path.join(os.path.dirname(file_name), var), variants_count, variants)
        else:
            print('Procesuję wybrany wariant:', os.path.basename(file_name))
            file_processing(file_name)
    else:
        file_processing(file_name)
    print(Fore.RESET + Back.RESET)


def variants_text(var, variants):
    """Generuje którki opis wariantu"""
    w = os.path.basename(var)
    try:
        variant = re.search(r'_[W|w]\d*_', w)[0].strip('_')
    except TypeError:
        variant = "W" + str(len(variants)+1)
    try:
        service = re.search(r'_SDI|_DLC|_PSK_', w, re.IGNORECASE)[0].strip('_').upper()
    except TypeError:
        service = "???"
    try:
        speed = re.search(r'_\d*M_', w)[0].strip('_')
    except TypeError:
        speed = "??Mbps"
    try:
        period = re.search(r'_\d*mc_', w)[0].strip('_')
    except TypeError:
        period = "??mc"
    # pattern = re.compile(r'(_SDI|_DLC|_PSK)(_W\d*_)(\d*M_)(\d*mc_)')
    # return pattern.search(w).group(2).strip('_') + ": " + pattern.search(w).group(1).strip('_') + " " + \
    #     pattern.search(w).group(3).strip('_') + " - umowa na okres " + pattern.search(w).group(4).strip('_')
    return variant + ": " + service + " " + speed + " - umowa na " + period

def main():
    init()
    filename = sys.argv[1]

    variants_check(filename)


main()
